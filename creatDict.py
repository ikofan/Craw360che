import openpyxl

# 建立字典的中文，截取单元格内容，只能截取短的内容，长的也可以，但是从翻译效果上可能不太好
def createDictList(excel_file):         # 读取excel表内容，取出不重复的词，建立待翻译的字典
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    list_set = set([])
    # zh_hans = re.compile(u'[\u4e00-\u9fa5]+')
    for column in ws["A:AA"]:
        cell_list = []       # 把整列单元格内容读入cellList中
        for cell in column:
            if is_contain_chinese(str(cell.value)):     # 此处有坑，一定要用str将单元格转成string格式，要不然报错
                cell_list.append(cell.value)
        list_set = list_set | set(cell_list)       # set()好像必须用列表类型才行，一开始用cell.value总是出错，于是加入cellList这个列表，解决
    dict_list = list(list_set)
    wb2 = openpyxl.Workbook()       #workbook一定要大写W，坑
    ws2 = wb2.active
    i = 1
    for dict in dict_list:
        ws2["A"+str(i)] = dict
        i += 1
    wb2.save("dict.xlsx")


def is_contain_chinese(check_str):
    """
    判断字符串中是否包含中文
    :param check_str: {str} 需要检测的字符串
    :return: {bool} 包含返回True， 不包含返回False
    """
    for ch in check_str:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


def is_chinese(string):
    """
    原理：
    中文字符的编码范围是：\u4e00 - \u9fff
    只要编码在此范围就可判断为中文字符
    检查整个字符串是否为中文
    Args:
        string (str): 需要检查的字符串,包含空格也是False
    Return
        bool
    """
    for chart in string:
        if chart < u'\u4e00' or chart > u'\u9fff':
            return False
    return True


def main():
    createDictList("result.xlsx")


if __name__ == '__main__':
    main()



# test("D:\\360che\\create_dict.xlsx")
