import openpyxl


wb = openpyxl.load_workbook("D:\\360che\\test\\dict.xlsx")
ws = wb.active

cn = []
en = []
for cell in ws["A"]:
    cn.append(str(cell.value))

for cell in ws["B"]:
    en.append(str(cell.value))

# zh_hans = tuple(cn)
# dict_list = dict.fromkeys(zh_hans)

dict_list = dict(zip(cn, en))
# 一开始还不会把列表转为字典，上面这个方法简单啊，学习了。

wb = openpyxl.load_workbook("D:\\360che\\test\\results_new_sorted.xlsx")
ws1 = wb["Sheet1"]
ws2 = wb.copy_worksheet(ws1)
ws2.title = 'EN'

# print(dict_list[t])

for col in ws2.columns:
    for cell in col:
        if str(cell.value) in dict_list:
            cell.value = dict_list[cell.value]


wb.save("D:\\360che\\test\\translate.xlsx")
