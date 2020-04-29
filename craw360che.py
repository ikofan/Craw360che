import re
import requests
from bs4 import BeautifulSoup
import xlwt, xlrd
import openpyxl
import lxml
import html5lib
import time

# 在写css时，标签名不加任何修饰，类名前加点，id名前加 #，我们可以用类似的方法来筛选元素，用到的方法是soup.select()，返回类型是list。


# 生成种子
def get_seeds():
    page = requests.get("http://www.360che.com/qianyinche/")
    soup = BeautifulSoup(page.content, "html5lib")
    # a_tags = soup.find_all(href=re.compile(".66_index.html"))
    # brand = soup.select('.xll_center2_a1_z dd')
    # for tag in a_tags:#  print(tag.get('href'), " ",tag.string)
    # print(soup.dt.string)
    div_tags = soup.select('.xll_center2_a1_y2 dt')
    source_urls = xlwt.Workbook(encoding='utf-8')
    ws = source_urls.add_sheet('urls')
    r = 0
    for tag in div_tags:
        ws.write(r, 0, tag.a.string)
        ws.write(r, 1, tag.a.get('href'))
        param = tag.a.get('href').replace('index', 'param')
        ws.write(r, 2, param)
        r += 1
    source_urls.save('urls.xls')


def get_summary(seed):
    page = requests.get(seed)
    soup = BeautifulSoup(page.content, 'html5lib')
    table = soup.select_one("table")
    para_summary = table.select("div .title-bar > h5")  # 所有在div标签 CSS class属性名为title-bar内的H5标签
    results = ["Summary"]
    for summary in para_summary:
        print("Get summary {}".format(summary.a.string))
        results.append(summary.a.string)
    print('Finished getting summary')
    return results
# 用于获取table标签下，thread-tr-th-div title-bar-h5-a下面标题的内容, 返回seed页面的标题列表
# seed格式为"https://product.360che.com/s25/6488_66_param.html"


def get_price(seed):
    # seed = "https://product.360che.com/s25/6488_66_param.html"
    page = requests.get(seed)
    soup = BeautifulSoup(page.content, 'html5lib')
    table = soup.select_one("table")    # 对应html中的第一个table标签
    price_tags = table.select('thead > tr')[1].select('td')     # table中的thead标签内的td标签，其中第二个td标签与价格有关
    # print(price_tags[1].string)
    results = ["Price"]
    for price in price_tags[1:]:    # 删掉厂商指导价
        print('Geting price...{}'.format(price.string))
        results.append(price.string)
    print('Finished getting price')
    return results


def load_seeds():           # 从excel表格读取种子页面的url，返回一个要抓取列表
    seeds_wb = xlrd.open_workbook('urls.xls')
    seeds_ws = seeds_wb.sheet_by_name('urls')
    seeds = seeds_ws.col_values(2)
    return seeds

# get_table("https://product.360che.com/s25/6488_66_param.html")
# get_tb("https://product.360che.com/s27/6778_66_param.html")


def wb_creat():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    ws.cell(1, 1, 4444)
    ws.insert_cols(1, 1)

    wb.save("d:\\test.xlsx")


def get_tb(seed):
    requests.adapters.DEFAULT_RETRIES = 5       # 出了几次连接频繁的错误，于是百度了一下采用这个方案
    print("Set retries = ", requests.adapters.DEFAULT_RETRIES)
    s = requests.sessions.session()

    s.keep_alive = False
    print("Trying to get seeds")

    page = s.get(seed)
    print("Get seeds result is {}".format(page.status_code))

    soup = BeautifulSoup(page.content, 'html5lib')
    table = soup.select_one("table")
    tr_class = table.find_all('tr', class_='param-row')            # 搜索所有param-row的class，输出列表，tr_class就是包含了所有列
    # td_tags = tr_class[2].find_all('td')
    # div_tags = tr_class[50].find_all('div')
    # print(td_tags)
    # print(div_tags[0].find(text=True).strip())
    # print(len(tr_class))
    ws_name = soup.find("div", class_='detail-header').h1.a.string.strip()          # 做表的名字需要注意，可能爬的字符又特殊字符或者空格，所以运行出错，加上strip成功
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(ws_name, cell_overwrite_ok=True)      # 把overwrite设置为true，否则后面插入列总是出错
    xl_col = 0
    for tr in tr_class:
        td_tag = tr.find('td', id=re.compile("ai_p_\d"))            # td_tag是excel的行标题
        div_tags = tr.find_all('div')
        ws.write(0, xl_col, td_tag.string)
        i = 1
        for div in div_tags:
            div_text = div.find(text=True).strip()
            print('Writing-- {}'.format(div_text))
            ws.write(i, xl_col, div_text)
            i += 1
        xl_col += 1

    summary = get_summary(seed)
    prices = get_price(seed)
    # print(summary)
    # print(prices)
    i = 0
    for summ in summary:
        print('Writing summary {}'.format(summ))
        ws.write(i, xl_col, summ)
        i += 1

    i = 0
    for price in prices:
        print('Writing price {}'.format(price))
        ws.write(i, xl_col+1, price)
        i += 1
    print("All DONE, Writing to EXCEL file")
    wb.save('G:\\360che\\'+ws_name+'.xls')


def main():
    seeds = load_seeds()
    for seed in seeds:
        print('crawling {}'.format(seed))
        get_tb(seed)


if __name__ == "__main__":
    main()

